from datetime import date, timedelta, datetime
from django.db.models import F
from django.shortcuts import render
from django.http import HttpResponse
from django.utils.timezone import make_naive
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.utils.dateparse import parse_date
from django.core.exceptions import ValidationError
from django.contrib.auth.decorators import login_required
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin

from ..atenciones.models import Reason, Service, Headquarter
from .. organismos.models import Organism

# Validación de fechas
def validate_date(date_str):
    try:
        parsed_date = datetime.strptime(date_str, '%d/%m/%Y').date()
    except ValueError:
        raise ValidationError(f'Invalid date: {date_str}')
    return parsed_date

def filter_attentions(request):
    attentions = Service.objects.all()
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')
    organism_id = request.GET.get('organism_id', '')
    service_reason_id = request.GET.get('service_reason_id', '')
    headquarter_id = request.GET.get('headquarter_id', '')
    min_age = request.GET.get('min_age', '')
    max_age = request.GET.get('max_age', '')
    persons_age = request.GET.get('persons_age', '')

    # Validación de fechas
    if from_date:
        from_date = validate_date(from_date)
    if to_date:
        to_date = validate_date(to_date)

    # Filtros
    if from_date or to_date:
        if from_date and to_date:
            attentions = attentions.filter(service_date__range=[from_date, to_date])
        elif from_date:
            attentions = attentions.filter(service_date__gte=from_date)
        elif to_date:
            attentions = attentions.filter(service_date__lte=to_date)

    if organism_id:
        attentions = attentions.filter(organism_id=organism_id)

    if service_reason_id:
        attentions = attentions.filter(service_reason_id=service_reason_id)

    if headquarter_id:
        attentions = attentions.filter(service_headquarter=headquarter_id)

    # Filtrado por edad
    today = date.today()

    if min_age:
        min_birthdate = today.replace(year=today.year - int(min_age))
        attentions = attentions.filter(person_id__person_birthdate__lte=min_birthdate)

    if max_age:
        max_birthdate = today.replace(year=today.year - int(max_age) - 1) + timedelta(days=1)
        attentions = attentions.filter(person_id__person_birthdate__gte=max_birthdate)

    if persons_age:
        persons_age = int(persons_age)
        exact_date = today - timedelta(days=(persons_age + 1) * 365)  # Fecha para edad exacta
        attentions = attentions.filter(
            person_id__person_birthdate__gte=exact_date,
            person_id__person_birthdate__lt=today - timedelta(days=persons_age * 365)  # Para que sea exactamente la edad
        )

    return attentions

class StatisticsView(LoginRequiredMixin, TemplateView):
    template_name = 'estadisticas/index.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Parámetros de solicitud (si existen)
        search = self.request.GET.get('search', '')  # Parámetro para indicar si se presionó "Buscar"
        from_date = self.request.GET.get('from_date', '')
        to_date = self.request.GET.get('to_date', '')
        organism_id = self.request.GET.get('organism_id', '')
        service_reason_id = self.request.GET.get('service_reason_id', '')
        headquarter_id = self.request.GET.get('headquarter_id', '')
        age = self.request.GET.get('age', '')
        min_age = self.request.GET.get('min_age', '')
        max_age = self.request.GET.get('max_age', '')
        persons_age = self.request.GET.get('persons_age', '')

        # Inicializar variables de contexto
        attentions = Service.objects.none()
        total_records = 0
        show_results = False

        # Asignar solo si se realiza la búsqueda
        if search:
            attentions = filter_attentions(self.request)

            # Paginación
            paginator = Paginator(attentions, 5)  # Mostrar 5 atenciones por página
            page = self.request.GET.get('page')  # Obtener el número de página

            try:
                attentions = paginator.page(page)
            except PageNotAnInteger:
                attentions = paginator.page(1)
            except EmptyPage:
                attentions = paginator.page(paginator.num_pages)

            total_records = attentions.paginator.count
            show_results = True

        # Obtener listas para desplegables
        headquarters = Headquarter.objects.all()
        reasons = Reason.objects.all()
        organisms = Organism.objects.all()

        # Recuperar los objetos para los filtros seleccionados, si fueron enviados
        selected_reason = Reason.objects.filter(service_reason_id=service_reason_id).first() if service_reason_id else None
        selected_headquarter = Headquarter.objects.filter(headquarter_id=headquarter_id).first() if headquarter_id else None
        selected_organism = Organism.objects.filter(organism_id=organism_id).first() if organism_id else None

        # Añadir al contexto solo si los valores fueron enviados
        context.update({
            'attentions': attentions,
            'headquarters': headquarters,
            'organisms': organisms,
            'reasons': reasons,
            'total_records': total_records,
            'show_results': show_results,
            'from_date': from_date,
            'to_date': to_date,
            'service_reason_id': service_reason_id if service_reason_id else None,
            'headquarter_id': headquarter_id if headquarter_id else None,
            'organism_id': organism_id if organism_id else None,
            'selected_reason': selected_reason,
            'selected_headquarter': selected_headquarter,
            'selected_organism': selected_organism,
            'age': age,
            'min_age': min_age,
            'max_age': max_age,
            'persons_age': persons_age,
        })

        return context

@login_required
def export_to_excel(request):
    # Obtener y filtrar atenciones
    attentions = filter_attentions(request)
    filename = request.GET.get('filename', 'atenciones.xlsx')

    # Crear archivo Excel
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Atenciones'

    # Escribir encabezados
    headers = ['ID', 'Motivo del Servicio', 'Sede de Atención', 'Persona Atendida', 'Edad' ,'Organismo', 'Fecha del Servicio']
    worksheet.append(headers)
    
    column_widths = [10, 30, 30, 30, 20, 40, 40]
    for i, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Aplicar estilos
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center')
    for cell in worksheet[1]:
        cell.alignment = header_alignment
        cell.font = header_font

    # Escribir datos
    for attention in attentions.select_related('service_reason_id', 'person_id', 'organism_id', 'service_headquarter'):
        service_date = make_naive(attention.service_date) if attention.service_date else ''
        person = attention.person_id
        age = person.age()  # Calcular la edad usando el método age() del modelo Person
        
        worksheet.append([
            attention.service_id,
            attention.service_reason_id.service_reason,
            attention.service_headquarter.headquarter_name,
            f"{person.person_surname} {person.person_name}",
            age,  # Aquí agregamos la edad
            attention.organism_id.organism_name if attention.organism_id else 'N/A',
            service_date.strftime('%d-%m-%Y %H:%M:%S') if service_date else '',
        ])

    # Aplicar formato a las celdas de datos
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in worksheet.iter_rows(min_row=2, max_col=worksheet.max_column, max_row=worksheet.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    # Agregar fila de total
    total_records = attentions.count()
    worksheet.append([''] * (len(headers) - 1) + [f'Total de atenciones: {total_records}'])

    # Aplicar estilo a la fila total
    total_row = worksheet.max_row
    for cell in worksheet[total_row]:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)

    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response